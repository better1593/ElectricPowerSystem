import openpyxl
import pandas as pd
import json
def process_row(sheet_name, row, headers):
    # 在这里添加你的行处理逻辑
    # 例如，根据sheet_name和row内容进行不同的处理
    processed_row = {}
    position = [row[5], row[6], row[7]]
    processed_row['position'] = position
    for h,r in zip(headers,row):  # 从第一列到第十三列
        if h == None:
            continue
        if h=='x0' or h=='y0' or h=='z0':
            continue
        processed_row[h] = r
    return processed_row

def read_tower(file_path):
    wb = openpyxl.load_workbook(file_path)
    data = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        headers = rows[1]  # 第二行作为列名
        if sheet_name == "Info":
            for row in rows[2:]:  # 从第三行开始处理数据
                first = row[0]
                if first == "info":
                    processed_row = process_row(sheet_name, row, headers)
                    data["name"] = row[1]
            data[sheet_name] = processed_row

        if sheet_name == "Device":
            devices = []
            device_dict = {}
            name_dict ={}
            for row in rows[2:]:
                first = row[0]
                device_name = row[1]
                if first  == "device":
                    if device_name not in device_dict:
                        device_dict[device_name] = []
                    new_entry = {h: r for h, r in zip(headers[4:], row[4:])}
                    device_dict[device_name].append(new_entry)
                    name_dict[device_name] = [row[1],row[2]]
            for key in list(device_dict.keys()):
                if sheet_name not in data:
                    data[sheet_name] = []
                new = {"name": name_dict[key][0], "type": name_dict[key][1], "Lump": device_dict[key]}
                data[sheet_name].append(new)
        if sheet_name == "Ground":
            for row in rows[2:]:
                first = row[0]
                if first == "ground":
                    ground = {h: r for h, r in zip(headers[1:], row[1:])}
                    data["ground"] = ground
        if sheet_name == "Wire":
            wires = []
            tube_dict = {}
            for row in rows[2:]:
                first = row[0]
                if first == "wire":
                    if row[1] == "air":
                        pos1 = [row[5], row[6], row[7]]
                        pos2 = [row[8], row[9], row[10]]
                        wire = {h: r for h, r in zip(headers[1:5] + headers[11:]+("pos_1","pos_2"), row[1:5] + row[11:]+(pos1,pos2))}
                        wires.append(wire)
                    if row[1].split("_")[0] == "tube":
                        if row[1].split("_")[1] not in tube_dict:
                            tube_dict[row[1].split("_")[1]] = {"type":"tube"}
                        if row[1].split("_")[2] =="sheath":
                            pos1 = [row[5], row[6], row[7]]
                            pos2 = [row[8], row[9], row[10]]
                            sheath = {h: r for h, r in zip(headers[2:5] + headers[15:]+("pos_1","pos_2","r0","rs1","rs3","type", headers[11]), row[2:5] + row[15:]+(pos1,pos2,row[13],row[12],row[14],"sheath",row[11]))}
                            tube_dict[row[1].split("_")[1]]["sheath"] = sheath
                        if row[1].split("_")[2] == "core":
                            pos1 = [row[5], row[6], row[7]]
                            pos2 = [row[8], row[9], row[10]]
                            core = {h: r for h, r in zip(headers[2:5] + headers[15:]+("pos_1","pos_2","oft","rc","cita","type", headers[11]), row[2:5] + row[15:]+(pos1,pos2,row[13],row[12],row[14],"core",row[11]))}
                            if "core" not in tube_dict[row[1].split("_")[1]]:
                                tube_dict[row[1].split("_")[1]]["core"] = []
                            tube_dict[row[1].split("_")[1]]["core"].append(core)
            wires = wires + list(tube_dict.values())
            data[sheet_name] = wires
        if sheet_name == "Lump":
            lumps = []
            for row in rows[2:]:
                first = row[0]
                if first == "lump":
                    lump = {h: r for h, r in zip(headers[1:], row[1:])}
                    lumps.append(lump)
            data[sheet_name] = lumps


    return data

def read_cable(file_path):
    wb = openpyxl.load_workbook(file_path)
    data = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        processed_data = []
        rows = list(sheet.iter_rows(values_only=True))
        headers = rows[1]  # 第二行作为列名
        if sheet_name == "Cable":
            print("cable")

def read_OHL(file_path):
    wb = openpyxl.load_workbook(file_path)
    data = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        headers = rows[1]  # 第二行作为列名

        if sheet_name == "Info":
            for row in rows[2:]:
                if row[0] == "info":
                    pos1 = [row[9], row[10], row[11]]
                    pos2 = [row[14], row[15], row[16]]
                    info = {h: r for h, r in zip(headers[1:9] + headers[12:14]+("Tower_head_pos","Tower_tail_pos"), row[1:9] + row[12:14]+(pos1,pos2))}
                    data[sheet_name] = info
                    data["name"] = row[1]

        if sheet_name == "Wire":
            wires = []
            for row in rows[2:]:
                if row[0] == "wire":
                    pos1 = [row[5], row[6], row[7]]
                    pos2 = [row[8], row[9], row[10]]
                    wire = {h: r for h, r in zip(headers[1:5] + headers[11:]+("node1_pos","node2_pos"), row[1:5] + row[11:]+(pos1,pos2))}
                    wires.append(wire)
            data[sheet_name] = wires

        if sheet_name == "ground":
            for row in rows[2:]:
                first = row[0]
                if first == "ground":
                    ground = {h: r for h, r in zip(headers[1:], row[1:])}
                    data[sheet_name] = ground
    return data

def read_source(file_path):
    wb = openpyxl.load_workbook(file_path)
    data = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        headers = rows[1]  # 第二行作为列名
        if sheet_name == "Lightning":
            for row in rows[2:]:
                if row[0] == "lightning":
                    pos1 = [row[4], row[5], row[6]]
                    info = {h: r for h, r in zip(headers[1:4] + headers[7:]+("position",), row[1:4] + row[7:]+(pos1,))}
                    data[sheet_name] = info
        if sheet_name == "Stroke":
            strokes = []
            for row in rows[2:]:
                stroke = {h: r for h, r in zip(headers[1:], row[1:])}
                strokes.append(stroke)
            data[sheet_name] = strokes
    return data

# 读取并处理每个文件的数据
tower_file = '../input/case1/Tower_Input.xlsx'  # 替换为你的tower文件路径
#cable_file = '../input/case1/Cable_Input.xlsx'  # 替换为你的cable文件路径
ohl_file = '../input/case1/Overheadline_Input.xlsx'      # 替换为你的ohl文件路径
source_file = '../input/case1/Source_Input.xlsx'

tower_data = read_tower(tower_file)
#cable_data = read_cable(cable_file)
ohl_data = read_OHL(ohl_file)
source_data = read_source(source_file)
# 合并数据到一个字典中
combined_data = {
    "user_id":"02",
    'case_id':1,
    'Tower': [tower_data],
    #'Cable': [cable_data],
    'OHL': [ohl_data],
    'Source': [source_file]
}

# 将数据转换为JSON格式
json_data = json.dumps(combined_data, ensure_ascii=False, indent=4)

# 将JSON数据写入文件
json_file = 'output.json'  # 替换为你想要保存的JSON文件路径
with open(json_file, 'w', encoding='utf-8') as f:
    f.write(json_data)

print(f"Excel文件已成功转换为JSON文件：{json_file}")

